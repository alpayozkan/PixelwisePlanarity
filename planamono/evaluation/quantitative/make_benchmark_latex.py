"""Generate LaTeX tables (one per dataset) from the XLSX summaries produced
by ``make_benchmark_summary.py``.

Reads the XLSX with **openpyxl** so per-cell bold formatting (the
moge-vs-zeroplane winner annotation) is preserved and re-emitted as
``\\textbf{...}`` in the LaTeX output.

For each dataset present in the XLSX, emits one ``<dataset>.tex`` file
under ``<exp_dir>/latex/`` containing a self-contained
``\\begin{table}...\\end{table}`` block. Also writes ``all_tables.tex`` that
``\\input``s every per-dataset file, suitable for dropping into a paper.

Default column groups (mirrors the picked-summary's column set):

- **Segmentation** — RI ↑, VI ↓, SC ↑
- **Plane Depth (%) ↑** — per-plane recall at 0.05 / 0.10 / 0.60 m
- **Plane Normal (%) ↑** — per-plane recall at 5° / 10° / 30°
- **Errors ↓** — Hungarian-matched mean normal (°) and offset (m) error
- **Plane fit @1mm (%) ↑** — RANSAC P / R at τ = 1 mm
- **Plane fit @5mm (%) ↑** — RANSAC P / R at τ = 5 mm
- **Plane fit @10mm (%) ↑** — RANSAC P / R at τ = 10 mm

Percentage cells (P/R/F1, per-plane depth/normal recall, etc.) are emitted
with a single decimal (``92.3 ± 4.1``); native-units cells (errors in
degrees / meters, RI/SC/VI) keep two decimals. Scaling and rounding are
applied upstream by ``make_benchmark_summary.py`` — this script just
passes formatted strings through.

Columns not present in the XLSX are silently skipped (you can point this
script at either the picked or complete XLSX).

Required LaTeX packages
-----------------------
The output uses ``booktabs``, ``multirow``, and inline math (``$\\pm$``,
``$\\uparrow$``, ``$\\downarrow$``). Make sure your preamble has::

    \\usepackage{booktabs}
    \\usepackage{multirow}

Usage
-----
    python make_benchmark_latex.py \\
        --xlsx /cluster/scratch/.../eval/<exp>/summary_complete_picked.xlsx \\
        [--output_dir /cluster/scratch/.../eval/<exp>/latex]

If ``--output_dir`` is omitted, files land next to the XLSX in a
sibling ``latex/`` directory.

Requires ``openpyxl`` (``pip install openpyxl``).
"""
from __future__ import annotations

import argparse
import os
import re
from typing import Dict, List, Optional, Tuple


# ---------------------------------------------------------------------------
# Layout configuration
# ---------------------------------------------------------------------------

# Each group has a name (becomes the multicolumn header) and a list of
# (xlsx_metric_name, latex_column_label, optional_arrow). If the arrow is
# None, no per-column direction is emitted (the group label carries it).
GroupSpec = Dict[str, object]

DEFAULT_GROUPS: List[GroupSpec] = [
    {
        "name": "Segmentation",
        "metrics": [
            ("RI", "RI", "↑"),
            ("VI", "VI", "↓"),
            ("SC", "SC", "↑"),
        ],
    },
    {
        "name": r"Plane Depth (\%) $\uparrow$",
        "metrics": [
            ("per_plane_depth_005", "@5cm", None),
            ("per_plane_depth_01",  "@10cm", None),
            ("per_plane_depth_06",  "@60cm", None),
        ],
    },
    {
        "name": r"Plane Normal (\%) $\uparrow$",
        "metrics": [
            ("per_plane_normal_5",  r"@5$^\circ$",  None),
            ("per_plane_normal_10", r"@10$^\circ$", None),
            ("per_plane_normal_30", r"@30$^\circ$", None),
        ],
    },
    {
        "name": r"Errors $\downarrow$",
        "metrics": [
            ("mean_normal_error_deg", r"Normal ($^\circ$)", None),
            ("mean_offset_error_m",   r"Offset (m)",        None),
        ],
    },
    {
        "name": r"@1mm (\%) $\uparrow$",
        "metrics": [
            ("prec@0.1cm", "P",  None),
            ("rec@0.1cm",  "R",  None),
            ("f1@0.1cm",   "F1", None),
        ],
    },
    {
        "name": r"@5mm (\%) $\uparrow$",
        "metrics": [
            ("prec@0.5cm", "P",  None),
            ("rec@0.5cm",  "R",  None),
            ("f1@0.5cm",   "F1", None),
        ],
    },
    {
        "name": r"@10mm (\%) $\uparrow$",
        "metrics": [
            ("prec@1.0cm", "P",  None),
            ("rec@1.0cm",  "R",  None),
            ("f1@1.0cm",   "F1", None),
        ],
    },
]

# Method ordering and label aliases. Methods not in the order list are
# appended in their XLSX order.
METHOD_ORDER: List[str] = ["gt", "zeroplane", "moge"]
METHOD_ALIAS: Dict[str, str] = {
    "gt":        r"GT \textit{(upper bound)}",
    "zeroplane": "ZeroPlane",
    "moge":      r"\textbf{Ours}",
}

# Dataset display labels and preferred order in the combined table.
# Datasets not in DATASET_ORDER are appended at the end in alphabetical order.
DATASET_ALIAS: Dict[str, str] = {
    "scannetpp":   "ScanNet++",
    "nyuv2":       "NYU-v2",
    "sevenscenes": "7-Scenes",
}
DATASET_ORDER: List[str] = ["scannetpp", "nyuv2", "sevenscenes"]


# ---------------------------------------------------------------------------
# XLSX loader (preserves bold)
# ---------------------------------------------------------------------------

def _strip_header(name: str) -> str:
    """Strip the direction-arrow / percent suffix added by
    make_benchmark_summary.py so column names match the canonical metric
    keys used in DEFAULT_GROUPS."""
    if name is None:
        return ""
    out = str(name)
    out = re.sub(r"\s+(↑|↓)", "", out)
    out = re.sub(r"\s+\(%\)", "", out)
    return out.strip()


def load_xlsx(xlsx_path: str) -> Tuple[List[str], List[List[Dict]]]:
    """Return (col_names, rows) where each cell is ``{"value", "bold"}``.

    ``col_names`` is canonicalised: arrows / "(%)" stripped so callers can
    look up by raw metric key.
    """
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=False))
    if not rows:
        return [], []

    headers = [_strip_header(c.value) for c in rows[0]]
    data: List[List[Dict]] = []
    for r in rows[1:]:
        data.append([{
            "value": c.value,
            "bold": bool(c.font is not None and c.font.bold),
        } for c in r])
    return headers, data


# ---------------------------------------------------------------------------
# LaTeX escaping + cell rendering
# ---------------------------------------------------------------------------

# Order matters: escape LaTeX-special chars BEFORE inserting math-mode
# replacements (which themselves contain $). Otherwise the second pass would
# escape the $ we just inserted.
_LATEX_PRE_ESCAPE = [
    ("\\", r"\textbackslash{}"),  # rare in our cells, but be safe
    ("&", r"\&"),
    ("#", r"\#"),
    ("$", r"\$"),
    ("%", r"\%"),
    ("_", r"\_"),
    ("{", r"\{"),
    ("}", r"\}"),
    ("~", r"\textasciitilde{}"),
    ("^", r"\textasciicircum{}"),
]
_LATEX_POST_INSERT = [
    ("±", r"$\pm$"),
    ("↑", r"$\uparrow$"),
    ("↓", r"$\downarrow$"),
]


def latex_escape(s) -> str:
    if s is None:
        return "-"
    out = str(s)
    for src, dst in _LATEX_PRE_ESCAPE:
        out = out.replace(src, dst)
    for src, dst in _LATEX_POST_INSERT:
        out = out.replace(src, dst)
    return out


def render_cell(cell: Dict, bold_override: bool = False) -> str:
    """Render an XLSX cell to LaTeX text, wrapping in ``\\textbf`` if bold."""
    text = latex_escape(cell["value"])
    if cell["bold"] or bold_override:
        return r"\textbf{" + text + "}"
    return text


def _strip_std(cell_value) -> str:
    """Drop the '± std' part of a 'mean ± std' XLSX cell, returning just the
    mean as a string. No-op for cells without ±."""
    if cell_value is None:
        return ""
    s = str(cell_value)
    if "±" in s:
        return s.split("±", 1)[0].strip()
    return s


def render_cell_no_std(cell: Dict, bold_override: bool = False) -> str:
    """Like ``render_cell`` but strips the std component first — yields only
    the mean, optionally wrapped in ``\\textbf``."""
    text = latex_escape(_strip_std(cell["value"]))
    if cell["bold"] or bold_override:
        return r"\textbf{" + text + "}"
    return text


# ---------------------------------------------------------------------------
# Table builder
# ---------------------------------------------------------------------------

def _select_groups(headers: List[str], groups: List[GroupSpec]) -> List[GroupSpec]:
    """Drop metrics absent from the XLSX, then drop groups that go empty."""
    out: List[GroupSpec] = []
    for g in groups:
        present = [(m, lbl, arr) for (m, lbl, arr) in g["metrics"] if m in headers]
        if present:
            out.append({"name": g["name"], "metrics": present})
    return out


def _ordered_method_indices(rows: List[List[Dict]],
                            method_col: int,
                            order: List[str]) -> List[int]:
    """Return row indices ordered by ``order`` first, then unspecified methods
    in their original sequence."""
    by_method = {row[method_col]["value"]: i for i, row in enumerate(rows)}
    out: List[int] = []
    seen = set()
    for m in order:
        if m in by_method:
            out.append(by_method[m])
            seen.add(m)
    for m, idx in by_method.items():
        if m not in seen:
            out.append(idx)
    return out


def build_table_for_dataset(
    headers: List[str],
    dataset_rows: List[List[Dict]],
    dataset_label: str,
    dataset_key: str,
    groups: List[GroupSpec],
    caption: Optional[str] = None,
    label: Optional[str] = None,
) -> str:
    method_col = headers.index("method")

    selected = _select_groups(headers, groups)
    if not selected:
        return ""

    ordered_idx = _ordered_method_indices(dataset_rows, method_col, METHOD_ORDER)

    # Total metric columns (right of the Method column).
    n_metric_cols = sum(len(g["metrics"]) for g in selected)
    col_spec = "l" + "c" * n_metric_cols

    lines: List[str] = []
    lines.append(r"\begin{table}[t]")
    lines.append(r"\centering")
    eff_caption = caption or f"Plane segmentation results on {dataset_label}."
    eff_label = label or f"tab:bench_{dataset_key}"
    lines.append(rf"\caption{{{eff_caption}}}")
    lines.append(rf"\label{{{eff_label}}}")
    lines.append(rf"\begin{{tabular}}{{{col_spec}}}")
    lines.append(r"\toprule")

    # Top header: Method (multirow) + group multicolumns
    top_parts = [r"\multirow{2}{*}{Method}"]
    cmidrules: List[str] = []
    col_cursor = 2  # column 1 is Method
    for g in selected:
        n = len(g["metrics"])
        top_parts.append(rf"\multicolumn{{{n}}}{{c}}{{{g['name']}}}")
        if n >= 1:
            cmidrules.append(rf"\cmidrule(lr){{{col_cursor}-{col_cursor + n - 1}}}")
        col_cursor += n
    lines.append(" & ".join(top_parts) + r" \\")
    lines.append(" ".join(cmidrules))

    # Sub-header: per-metric short label + optional arrow
    sub_parts = [""]  # under the multirow Method
    for g in selected:
        for (_, label_text, arrow) in g["metrics"]:
            cell = label_text
            if arrow == "↑":
                cell = cell + r" $\uparrow$"
            elif arrow == "↓":
                cell = cell + r" $\downarrow$"
            sub_parts.append(cell)
    lines.append(" & ".join(sub_parts) + r" \\")
    lines.append(r"\midrule")

    # Data rows
    for row_idx in ordered_idx:
        row = dataset_rows[row_idx]
        method_raw = row[method_col]["value"]
        method_lbl = METHOD_ALIAS.get(method_raw, method_raw)
        cells: List[str] = [method_lbl]
        for g in selected:
            for (mname, _, _) in g["metrics"]:
                col_j = headers.index(mname)
                cells.append(render_cell(row[col_j]))
        lines.append(" & ".join(cells) + r" \\")

    lines.append(r"\bottomrule")
    lines.append(r"\end{tabular}")
    lines.append(r"\end{table}")
    return "\n".join(lines)


def _render_combined_body(
    headers: List[str],
    rows_by_dataset: Dict[str, List[List[Dict]]],
    selected_groups: List[GroupSpec],
    dataset_order: List[str],
    method_aliases: Dict[str, str],
    cell_fn,
) -> List[str]:
    """Shared core: header + per-dataset row blocks, between ``\\toprule``
    and ``\\bottomrule``. Caller wraps with the table preamble/postamble.
    """
    method_col = headers.index("method")

    known = [d for d in dataset_order if d in rows_by_dataset]
    extras = sorted(d for d in rows_by_dataset.keys() if d not in known)
    ordered_datasets = known + extras

    lines: List[str] = []

    # Top header.
    top_parts = [r"\multirow{2}{*}{Dataset}", r"\multirow{2}{*}{Method}"]
    cmidrules: List[str] = []
    col_cursor = 3
    for g in selected_groups:
        n = len(g["metrics"])
        top_parts.append(rf"\multicolumn{{{n}}}{{c}}{{{g['name']}}}")
        cmidrules.append(rf"\cmidrule(lr){{{col_cursor}-{col_cursor + n - 1}}}")
        col_cursor += n
    lines.append(" & ".join(top_parts) + r" \\")
    lines.append(" ".join(cmidrules))

    # Sub-header.
    sub_parts = ["", ""]
    for g in selected_groups:
        for (_, label_text, arrow) in g["metrics"]:
            cell = label_text
            if arrow == "↑":
                cell = cell + r" $\uparrow$"
            elif arrow == "↓":
                cell = cell + r" $\downarrow$"
            sub_parts.append(cell)
    lines.append(" & ".join(sub_parts) + r" \\")
    lines.append(r"\midrule")

    for di, ds in enumerate(ordered_datasets):
        ds_rows = rows_by_dataset[ds]
        ordered_idx = _ordered_method_indices(ds_rows, method_col, METHOD_ORDER)
        ds_label = DATASET_ALIAS.get(ds, ds)

        if di > 0:
            lines.append(r"\midrule")

        for ri, row_idx in enumerate(ordered_idx):
            row = ds_rows[row_idx]
            method_raw = row[method_col]["value"]
            method_lbl = method_aliases.get(method_raw, method_raw)

            if ri == 0:
                ds_cell = rf"\multirow{{{len(ordered_idx)}}}{{*}}{{{ds_label}}}"
            else:
                ds_cell = ""

            cells: List[str] = [ds_cell, method_lbl]
            for g in selected_groups:
                for (mname, _, _) in g["metrics"]:
                    col_j = headers.index(mname)
                    cells.append(cell_fn(row[col_j]))
            lines.append(" & ".join(cells) + r" \\")

    return lines


def build_combined_table(
    headers: List[str],
    rows_by_dataset: Dict[str, List[List[Dict]]],
    groups: List[GroupSpec],
    dataset_order: List[str],
    caption: str = "Plane segmentation results across datasets.",
    label: str = "tab:bench_all",
) -> str:
    """One table with Dataset as the leading column. ``\\midrule`` separates
    each dataset block; the dataset cell uses ``\\multirow`` to span its
    methods. Cells show ``mean ± std``.
    """
    selected = _select_groups(headers, groups)
    if not selected:
        return ""

    n_metric_cols = sum(len(g["metrics"]) for g in selected)
    col_spec = "ll" + "c" * n_metric_cols

    body = _render_combined_body(
        headers=headers,
        rows_by_dataset=rows_by_dataset,
        selected_groups=selected,
        dataset_order=dataset_order,
        method_aliases=METHOD_ALIAS,
        cell_fn=render_cell,
    )

    lines: List[str] = []
    lines.append(r"\begin{table*}[t]")
    lines.append(r"\centering")
    lines.append(rf"\caption{{{caption}}}")
    lines.append(rf"\label{{{label}}}")
    lines.append(rf"\begin{{tabular}}{{{col_spec}}}")
    lines.append(r"\toprule")
    lines.extend(body)
    lines.append(r"\bottomrule")
    lines.append(r"\end{tabular}")
    lines.append(r"\end{table*}")
    return "\n".join(lines)


# Compact variant: drops std, uses tiny font + tight spacing + resizebox.
# Same data, optimised for paper "main results" tables.
COMPACT_METHOD_ALIAS: Dict[str, str] = {
    "gt":        r"GT \textit{(UB)}",
    "zeroplane": "ZeroPlane",
    "moge":      r"\textbf{Ours}",
}

COMPACT_DEFAULT_CAPTION = (
    r"Plane segmentation results across datasets (ScanNet++, NYU-v2, 7-Scenes). "
    r"Ours achieves superior geometric accuracy (lower normal error, offset error) "
    r"and better segmentation quality (lower VI score) while maintaining "
    r"competitive recall metrics across all datasets."
)


def build_combined_table_compact(
    headers: List[str],
    rows_by_dataset: Dict[str, List[List[Dict]]],
    groups: List[GroupSpec],
    dataset_order: List[str],
    caption: str = COMPACT_DEFAULT_CAPTION,
    label: str = "tab:bench_all_compact",
) -> str:
    """Compact (mean-only, tiny-font, resizeboxed) variant of the combined
    table. No std, ``GT (UB)`` instead of ``GT (upper bound)``.
    """
    selected = _select_groups(headers, groups)
    if not selected:
        return ""

    n_metric_cols = sum(len(g["metrics"]) for g in selected)
    col_spec = "ll" + "c" * n_metric_cols

    body = _render_combined_body(
        headers=headers,
        rows_by_dataset=rows_by_dataset,
        selected_groups=selected,
        dataset_order=dataset_order,
        method_aliases=COMPACT_METHOD_ALIAS,
        cell_fn=render_cell_no_std,
    )

    lines: List[str] = []
    lines.append(r"\begin{table*}[t]")
    lines.append(r"\centering")
    lines.append(rf"\caption{{{caption}}}")
    lines.append(rf"\label{{{label}}}")
    lines.append(r"\tiny")
    lines.append(r"\setlength{\tabcolsep}{0.8pt}      % ultra tight spacing")
    lines.append(r"\renewcommand{\arraystretch}{0.65}  % very compact rows")
    lines.append(r"\resizebox{\textwidth}{!}{%")
    lines.append(rf"\begin{{tabular}}{{{col_spec}}}")
    lines.append(r"\toprule")
    lines.extend(body)
    lines.append(r"\bottomrule")
    lines.append(r"\end{tabular}}")
    lines.append(r"\end{table*}")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Driver
# ---------------------------------------------------------------------------

DEFAULT_XLSX = (
    "/cluster/scratch/aoezkan/planeseg/eval/"
    "gt_moge_zp_benchmark_kunscaled_old/summary_complete_picked.xlsx"
)


def main():
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("--xlsx", default=DEFAULT_XLSX,
                    help=f"Path to summary XLSX "
                         f"(default: {DEFAULT_XLSX}).")
    ap.add_argument("--output_dir", default=None,
                    help="Where to write per-dataset .tex files "
                         "(default: <xlsx_dir>/latex/).")
    ap.add_argument("--datasets", nargs="*", default=None,
                    help="Optional subset of datasets (matches the 'dataset' "
                         "column verbatim). Default: all.")
    args = ap.parse_args()

    if not os.path.isfile(args.xlsx):
        raise SystemExit(f"xlsx not found: {args.xlsx}")

    output_dir = args.output_dir or os.path.join(
        os.path.dirname(os.path.abspath(args.xlsx)), "latex"
    )
    os.makedirs(output_dir, exist_ok=True)

    headers, rows = load_xlsx(args.xlsx)
    if "method" not in headers or "dataset" not in headers:
        raise SystemExit("xlsx must contain 'method' and 'dataset' columns")

    method_col = headers.index("method")
    ds_col = headers.index("dataset")

    # Group rows by dataset.
    by_dataset: Dict[str, List[List[Dict]]] = {}
    for row in rows:
        ds = row[ds_col]["value"]
        by_dataset.setdefault(ds, []).append(row)

    targets = args.datasets or sorted(by_dataset.keys())
    written: List[str] = []
    for ds in targets:
        if ds not in by_dataset:
            print(f"  [skip] dataset '{ds}' not in XLSX")
            continue
        ds_label = DATASET_ALIAS.get(ds, ds)
        tex = build_table_for_dataset(
            headers=headers,
            dataset_rows=by_dataset[ds],
            dataset_label=ds_label,
            dataset_key=ds,
            groups=DEFAULT_GROUPS,
        )
        if not tex:
            print(f"  [skip] {ds}: no matching metrics in XLSX")
            continue
        out_path = os.path.join(output_dir, f"{ds}.tex")
        with open(out_path, "w") as f:
            f.write(tex + "\n")
        written.append(ds)
        print(f"  wrote {out_path}")

    # Combined wrapper that \input's every per-dataset table.
    if written:
        combined_path = os.path.join(output_dir, "all_tables.tex")
        with open(combined_path, "w") as f:
            f.write("% Auto-generated by make_benchmark_latex.py\n")
            f.write("% Required: \\usepackage{booktabs}, \\usepackage{multirow}\n\n")
            for ds in written:
                f.write(rf"\input{{{ds}.tex}}" + "\n")
        print(f"  wrote {combined_path}")

    # Single combined table (Dataset column + \midrule between datasets).
    if by_dataset:
        # Filter to requested datasets if --datasets was given.
        scope = {ds: by_dataset[ds] for ds in (args.datasets or by_dataset)
                 if ds in by_dataset}
        combined_one = build_combined_table(
            headers=headers,
            rows_by_dataset=scope,
            groups=DEFAULT_GROUPS,
            dataset_order=DATASET_ORDER,
        )
        if combined_one:
            single_path = os.path.join(output_dir, "all_combined.tex")
            with open(single_path, "w") as f:
                f.write(combined_one + "\n")
            print(f"  wrote {single_path}")

        # Compact variant: same content, no std, tiny font + resizebox.
        combined_compact = build_combined_table_compact(
            headers=headers,
            rows_by_dataset=scope,
            groups=DEFAULT_GROUPS,
            dataset_order=DATASET_ORDER,
        )
        if combined_compact:
            compact_path = os.path.join(output_dir, "all_combined_compact.tex")
            with open(compact_path, "w") as f:
                f.write(combined_compact + "\n")
            print(f"  wrote {compact_path}")


if __name__ == "__main__":
    main()
