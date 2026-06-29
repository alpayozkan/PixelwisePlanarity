"""Build pooled-per-frame XLSX summaries from a benchmark experiment dir.

Reads ``<exp_dir>/<method>/<dataset>/aggregate_results.csv`` (one per
(method, dataset) cell — concatenation of every per-frame row) produced by
``evaluate_gt_moge_zeroplane_benchmark.py``, then computes per-(method,
dataset) statistics by **pooling across all frames in the dataset** (not
across scene-means). This gives a meaningful std for single-scene datasets
like NYU-v2, which the existing per-scene aggregator leaves as NaN.

Outputs two XLSX files in the experiment directory:

  1. ``summary_complete.xlsx``       — all metric columns from the source CSVs.
  2. ``summary_complete_picked.xlsx`` — a curated subset (RI/VI/SC, plane-
     level depth/normal recall, Hungarian normal/offset error, plus the
     planamono RANSAC prec/rec block).

Both files format each metric cell as ``mean ± std`` and **bold the winner**
between ``moge`` and ``zeroplane`` per (dataset, metric) row, matching
``prettify_benchmark_summary.py``'s convention. GT is unbolded (it's the
upper bound, not in the comparison).

Pooling note: for each (method, dataset), mean = ``df[metric].mean()`` over
every per-frame row in that dataset; std = ``.std()`` (ddof=1) over the
same rows. So ScanNet++ pools over ~7000+ frames, NYU-v2 over 654, 7-Scenes
over ~6000 (depending on the split). Scene structure is ignored at this
stage — use the existing per-scene/per-dataset CSVs for scene-level views.

Usage::

    python make_benchmark_summary.py \\
        --exp_dir /cluster/scratch/aoezkan/planeseg/eval/gt_moge_zp_benchmark_kunscaled_old

Both XLSX files land in ``--exp_dir``. Override paths with
``--complete_out`` / ``--picked_out`` if needed.

Requires ``openpyxl`` (``pip install openpyxl``).
"""
from __future__ import annotations

import argparse
import math
import os
import re
from typing import Dict, List, Optional, Tuple

import pandas as pd


# ---------------------------------------------------------------------------
# Metric direction (re-used from prettify_benchmark_summary.py)
# ---------------------------------------------------------------------------

LOWER_IS_BETTER_PATTERNS: List[re.Pattern] = [
    re.compile(r"^VI$"),
    re.compile(r"^DE_rel$"), re.compile(r"^DE_rel_sqr$"), re.compile(r"^DE_log10$"),
    re.compile(r"^DE_rmse$"), re.compile(r"^DE_rmse_log$"),
    re.compile(r"^pixel_DE_rel.*$"),
    re.compile(r"^pixel_DE_log10$"), re.compile(r"^pixel_DE_rmse.*$"),
    re.compile(r"^mean_normal_error_deg$"), re.compile(r"^median_normal_error_deg$"),
    re.compile(r"^mean_offset_error_m$"), re.compile(r"^median_offset_error_m$"),
    re.compile(r"^plane_param_L2_mean$"),
    re.compile(r"^plane_param_L2_area_weighted$"),
]
HIGHER_IS_BETTER_PATTERNS: List[re.Pattern] = [
    re.compile(r"^RI$"), re.compile(r"^SC$"),
    re.compile(r"^DE_accuracy_[123]$"), re.compile(r"^pixel_DE_accuracy_[123]$"),
    re.compile(r"^per_pixel_depth_.*$"), re.compile(r"^per_plane_depth_.*$"),
    re.compile(r"^per_pixel_normal_.*$"), re.compile(r"^per_plane_normal_.*$"),
    re.compile(r"^per_pixel_offset_.*$"), re.compile(r"^per_plane_offset_.*$"),
    re.compile(r"^AP@\d+cm$"),
    re.compile(r"^prec@.*cm$"), re.compile(r"^rec@.*cm$"), re.compile(r"^f1@.*cm$"),
]


def metric_direction(metric: str) -> Optional[str]:
    for p in LOWER_IS_BETTER_PATTERNS:
        if p.match(metric):
            return "↓"
    for p in HIGHER_IS_BETTER_PATTERNS:
        if p.match(metric):
            return "↑"
    return None


# ---------------------------------------------------------------------------
# The curated picked-metric set
# ---------------------------------------------------------------------------

PICKED_METRICS: List[str] = [
    "RI", "VI", "SC",
    "per_plane_depth_005", "per_plane_depth_01", "per_plane_depth_06",
    "per_plane_normal_5", "per_plane_normal_10", "per_plane_normal_30",
    "mean_normal_error_deg",     # CRITICAL
    "mean_offset_error_m",       # CRITICAL
    "prec@0.1cm", "rec@0.1cm", "f1@0.1cm",
    "prec@0.5cm", "rec@0.5cm", "f1@0.5cm",
    "prec@1.0cm", "rec@1.0cm", "f1@1.0cm",
]

# Columns that are not metrics — never aggregated, kept as-is in row labels.
LABEL_COLS = ("scene_id", "frame_idx")

# Metric families displayed as percentages (×100). The underlying values are
# already in [0, 1]; we multiply mean and std by 100 in the rendered cell and
# tag the header with "(%)". Kept in sync with prettify_benchmark_summary.py.
#
# Currently includes: RANSAC plane-fit P/R/F1, per-{pixel,plane} recall
# (depth/normal/offset), δ<1.25^n depth-accuracy thresholds, and AP@<τ>cm.
# All are bounded in [0, 1] by construction. Unbounded errors (VI, depth
# RMSE, normal-error degrees, offset-error meters) stay in their native units.
PCT_METRIC_PATTERNS: List[re.Pattern] = [
    re.compile(r"^prec@.*cm$"),
    re.compile(r"^rec@.*cm$"),
    re.compile(r"^f1@.*cm$"),
    re.compile(r"^per_pixel_depth_.*$"),
    re.compile(r"^per_plane_depth_.*$"),
    re.compile(r"^per_pixel_normal_.*$"),
    re.compile(r"^per_plane_normal_.*$"),
    re.compile(r"^per_pixel_offset_.*$"),
    re.compile(r"^per_plane_offset_.*$"),
    re.compile(r"^DE_accuracy_[123]$"),
    re.compile(r"^pixel_DE_accuracy_[123]$"),
    re.compile(r"^AP@\d+cm$"),
]


def _is_pct_metric(metric: str) -> bool:
    return any(p.match(metric) for p in PCT_METRIC_PATTERNS)


# Map prec@<τ>cm → its matching rec@<τ>cm so we can synthesise f1@<τ>cm
# per-frame. We do this BEFORE aggregating so std(F1) is computed across
# frames, not derived (incorrectly) from std(P) and std(R).
_PREC_REC_RE = re.compile(r"^prec@(.*)cm$")


def _add_f1_columns(df) -> None:
    """In-place: for each prec@<τ>cm with a matching rec@<τ>cm, add a
    f1@<τ>cm column computed per-row as the harmonic mean.
    """
    import numpy as np
    for c in list(df.columns):
        m = _PREC_REC_RE.match(c)
        if not m:
            continue
        tag = m.group(1)
        rec_col = f"rec@{tag}cm"
        if rec_col not in df.columns:
            continue
        f1_col = f"f1@{tag}cm"
        if f1_col in df.columns:
            continue   # already present, don't overwrite
        p = df[c].astype(float)
        r = df[rec_col].astype(float)
        denom = p + r
        # Avoid 0/0; F1 is 0 if both P and R are 0, NaN if either is NaN.
        with np.errstate(invalid="ignore", divide="ignore"):
            f1 = np.where(denom > 0, 2.0 * p * r / denom, 0.0)
        # Propagate NaN where either input is NaN.
        f1 = np.where(p.isna() | r.isna(), np.nan, f1)
        df[f1_col] = f1


# ---------------------------------------------------------------------------
# Pooled aggregation
# ---------------------------------------------------------------------------

def _read_per_frame_rows(dataset_dir: str) -> Optional[pd.DataFrame]:
    """Prefer ``aggregate_results.csv`` (single concat). Fall back to
    walking ``<scene>/results.csv`` if the aggregator hasn't run yet.
    """
    agg_path = os.path.join(dataset_dir, "aggregate_results.csv")
    if os.path.isfile(agg_path):
        try:
            return pd.read_csv(agg_path)
        except Exception as e:
            print(f"  [warn] failed to read {agg_path}: {e}")
    # Fallback: concat per-scene results.csv.
    parts: List[pd.DataFrame] = []
    for entry in sorted(os.listdir(dataset_dir)):
        rcsv = os.path.join(dataset_dir, entry, "results.csv")
        if os.path.isfile(rcsv):
            try:
                parts.append(pd.read_csv(rcsv))
            except Exception as e:
                print(f"  [warn] failed to read {rcsv}: {e}")
    if not parts:
        return None
    return pd.concat(parts, ignore_index=True)


def collect_pooled(exp_dir: str) -> Tuple[pd.DataFrame, List[str]]:
    """Walk <exp_dir>/<method>/<dataset>/ and return one row per
    (method, dataset) with pooled mean/std for every numeric metric column.

    Returns (rows_df, metric_names). Each metric ``M`` produces two columns
    ``M_mean`` and ``M_std`` in the output frame.
    """
    rows: List[Dict] = []
    metric_set: set = set()

    methods = sorted(d for d in os.listdir(exp_dir)
                     if os.path.isdir(os.path.join(exp_dir, d)))
    for method in methods:
        m_dir = os.path.join(exp_dir, method)
        for dataset in sorted(os.listdir(m_dir)):
            d_dir = os.path.join(m_dir, dataset)
            if not os.path.isdir(d_dir):
                continue
            df = _read_per_frame_rows(d_dir)
            if df is None or df.empty:
                continue

            # Synthesise f1@<τ>cm per-frame so std(F1) is the across-frame
            # standard deviation (not a function of std(P), std(R)).
            _add_f1_columns(df)

            row: Dict[str, float] = {
                "method": method,
                "dataset": dataset,
                "num_frames": int(len(df)),
            }
            for c in df.columns:
                if c in LABEL_COLS:
                    continue
                if df[c].dtype.kind not in ("f", "i"):
                    continue
                vals = df[c].dropna()
                if vals.empty:
                    row[f"{c}_mean"] = float("nan")
                    row[f"{c}_std"] = float("nan")
                else:
                    row[f"{c}_mean"] = float(vals.mean())
                    row[f"{c}_std"] = float(vals.std()) if len(vals) > 1 else float("nan")
                metric_set.add(c)
            rows.append(row)

    return pd.DataFrame(rows), sorted(metric_set)


# ---------------------------------------------------------------------------
# Cell formatting + bolding
# ---------------------------------------------------------------------------

def _fmt_mean_std(mean: float, std: float, ndigits: int = 2,
                  scale: float = 1.0) -> str:
    """Format ``mean ± std`` with optional ``×scale`` multiplier (for pct
    metrics). NaN std → mean only; NaN mean → empty cell."""
    if mean is None or (isinstance(mean, float) and math.isnan(mean)):
        return ""
    m = mean * scale
    if std is None or (isinstance(std, float) and math.isnan(std)):
        return f"{m:.{ndigits}f}"
    s = std * scale
    return f"{m:.{ndigits}f} ± {s:.{ndigits}f}"


def _display_round(v: float, is_pct: bool) -> float:
    """Round a raw [0, 1] (or native-units) value to the precision the cell
    will be rendered with. Used by the tie-detection in ``build_table`` so
    two cells that print the same string both get bolded.

    Pct metrics render as ``mean*100`` with 1 decimal → 0.1% absolute
    resolution → equivalent to rounding the raw value to 3 decimals.
    Native-units cells render with 2 decimals.
    """
    if is_pct:
        return round(v * 100.0, 1) / 100.0
    return round(v, 2)


def build_table(
    rows_df: pd.DataFrame,
    metrics: List[str],
) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, str]]:
    """Returns (display_df, bold_mask, header_renames).

    - ``display_df``: one row per (method, dataset). Metric cells are
      ``"mean ± std"`` strings; label cells are raw values.
    - ``bold_mask``: same shape, True where the cell should render bold.
      Applies to method/dataset label columns (anchor) and to the winner
      between moge/zeroplane per (dataset, metric).
    - ``header_renames``: original col → display col with ↑ / ↓.
    """
    label_cols = ["method", "dataset", "num_frames"]
    out_cols = label_cols + list(metrics)

    out_rows = []
    for _, row in rows_df.iterrows():
        out_row = {c: row.get(c, "") for c in label_cols}
        for m in metrics:
            mean = row.get(f"{m}_mean", float("nan"))
            std = row.get(f"{m}_std", float("nan"))
            # Percentages get 1 decimal (saves table width); native-units
            # metrics keep 2 decimals.
            if _is_pct_metric(m):
                scale, ndigits = 100.0, 1
            else:
                scale, ndigits = 1.0, 2
            out_row[m] = _fmt_mean_std(mean, std, ndigits=ndigits, scale=scale)
        out_rows.append(out_row)
    display_df = pd.DataFrame(out_rows, columns=out_cols)

    bold = pd.DataFrame(False, index=display_df.index, columns=out_cols)
    for c in ("method", "dataset"):
        bold[c] = True

    # Bold the winner(s) across {moge, zeroplane, metric3d} per (dataset,
    # metric). gt is excluded as the upper bound; moge_ep2 is excluded by
    # design (it's a sibling ablation, not in the headline comparison).
    # Ties bold every row that matches the optimal value, where "tie" is
    # measured at the same precision the cell will be rendered with — so
    # two cells that print the same string both get bolded.
    BOLD_METHODS = {"moge", "zeroplane", "metric3d"}
    for dataset, group in rows_df.groupby("dataset"):
        cand_rows = group[group["method"].isin(BOLD_METHODS)]
        if len(cand_rows) < 2:
            continue
        for m in metrics:
            d = metric_direction(m)
            if d is None:
                continue
            mean_col = f"{m}_mean"
            if mean_col not in rows_df.columns:
                continue
            is_pct = _is_pct_metric(m)
            vals_disp: Dict[int, float] = {}
            for ri in cand_rows.index:
                try:
                    v = float(rows_df.at[ri, mean_col])
                except (TypeError, ValueError):
                    continue
                if not math.isnan(v):
                    vals_disp[int(ri)] = _display_round(v, is_pct)
            if len(vals_disp) < 2:
                continue
            best = max(vals_disp.values()) if d == "↑" else min(vals_disp.values())
            for ri, vr in vals_disp.items():
                if vr == best:
                    bold.at[ri, m] = True

    renames = {c: c for c in label_cols}
    for m in metrics:
        d = metric_direction(m) or ""
        suffix = " (%)" if _is_pct_metric(m) else ""
        renames[m] = f"{m} {d}{suffix}".strip()

    return display_df, bold, renames


# ---------------------------------------------------------------------------
# XLSX writer
# ---------------------------------------------------------------------------

def write_xlsx(
    display_df: pd.DataFrame,
    bold: pd.DataFrame,
    renames: Dict[str, str],
    out_path: str,
    sheet_title: str = "summary",
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise SystemExit(
            "openpyxl is required for xlsx output. Install with:\n"
            "    pip install openpyxl"
        ) from e

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    bold_font = Font(bold=True)

    cols = list(display_df.columns)
    headers = [renames.get(c, c) for c in cols]
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.font = bold_font

    for i, (_, row) in enumerate(display_df.iterrows(), start=2):
        for j, c in enumerate(cols, start=1):
            v = row[c]
            cell = ws.cell(row=i, column=j, value=(v if v != "" else None))
            if bool(bold.iat[i - 2, j - 1]):
                cell.font = bold_font

    # Freeze header + the (method, dataset) anchor columns.
    ws.freeze_panes = "C2"

    # Reasonable column widths.
    for j, c in enumerate(cols, start=1):
        col_letter = get_column_letter(j)
        header_len = len(str(headers[j - 1]))
        sample_len = display_df[c].astype(str).str.len().max() if len(display_df) else 0
        try:
            sample_len = int(sample_len)
        except (TypeError, ValueError):
            sample_len = 0
        ws.column_dimensions[col_letter].width = min(40, max(8, header_len + 2, sample_len + 2))

    wb.save(out_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

DEFAULT_EXP_DIR = "/cluster/scratch/aoezkan/planeseg/eval/gt_moge_zp_benchmark_kunscaled_old"


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--exp_dir", default=DEFAULT_EXP_DIR,
                    help=f"Experiment dir (default: {DEFAULT_EXP_DIR}).")
    ap.add_argument("--complete_out", default=None,
                    help="Override path for summary_complete.xlsx "
                         "(default: <exp_dir>/summary_complete.xlsx).")
    ap.add_argument("--picked_out", default=None,
                    help="Override path for summary_complete_picked.xlsx "
                         "(default: <exp_dir>/summary_complete_picked.xlsx).")
    args = ap.parse_args()

    if not os.path.isdir(args.exp_dir):
        raise SystemExit(f"exp_dir not found: {args.exp_dir}")

    rows_df, all_metrics = collect_pooled(args.exp_dir)
    if rows_df.empty:
        raise SystemExit(f"No per-frame data found under {args.exp_dir}")

    print(f"[summary] {len(rows_df)} (method, dataset) cells, "
          f"{len(all_metrics)} metric columns")
    for _, r in rows_df[["method", "dataset", "num_frames"]].iterrows():
        print(f"  {r['method']:<10s} / {r['dataset']:<11s}  "
              f"{int(r['num_frames']):>5d} frames")

    # Complete
    complete_out = args.complete_out or os.path.join(args.exp_dir, "summary_complete.xlsx")
    display_df, bold, renames = build_table(rows_df, all_metrics)
    write_xlsx(display_df, bold, renames, complete_out, sheet_title="all_metrics")
    print(f"[summary] wrote {complete_out}  "
          f"({len(display_df)} rows × {len(display_df.columns)} cols)")

    # Picked
    picked_metrics = [m for m in PICKED_METRICS if m in all_metrics]
    missing = [m for m in PICKED_METRICS if m not in all_metrics]
    if missing:
        print(f"[summary]  warning: picked metrics not present in data: {missing}")
    picked_out = args.picked_out or os.path.join(args.exp_dir, "summary_complete_picked.xlsx")
    display_df_p, bold_p, renames_p = build_table(rows_df, picked_metrics)
    write_xlsx(display_df_p, bold_p, renames_p, picked_out, sheet_title="picked")
    print(f"[summary] wrote {picked_out}  "
          f"({len(display_df_p)} rows × {len(display_df_p.columns)} cols)")

    # Direction-map gap warnings (mirrors prettify_benchmark_summary.py).
    unknown = [m for m in all_metrics if metric_direction(m) is None]
    if unknown:
        print(f"[summary]  no direction set for {len(unknown)} metric(s) "
              f"(no bolding for these): {unknown}")


if __name__ == "__main__":
    main()
