"""Prettify ``summary.csv`` from ``evaluate_gt_moge_zeroplane_benchmark.py``
into an XLSX with real bold cells.

- Drops every ``*_std`` column; keeps only the ``*_mean`` value of each metric.
- Renames metric columns: strips the ``_mean`` suffix and appends ↑ / ↓ to
  indicate direction (higher / lower is better).
- For each dataset row, compares moge vs zeroplane per metric and applies
  ``Font(bold=True)`` to the winner cell. GT is left unbolded (it's the
  upper bound, not in the comparison).
- The header row and the ``method`` / ``dataset`` columns are also bold.

Usage
-----
    python prettify_benchmark_summary.py \\
        --input  /cluster/scratch/aoezkan/planeseg/eval/gt_moge_zp_benchmark/summary.csv \\
        --output /cluster/scratch/aoezkan/planeseg/eval/gt_moge_zp_benchmark/summary_pretty.xlsx

Defaults point at the standard benchmark output if both flags are omitted.

Requires ``openpyxl`` (``pip install openpyxl``).
"""
from __future__ import annotations

import argparse
import math
import os
import re
from typing import Dict, List, Optional

import pandas as pd


DEFAULT_INPUT = "/cluster/scratch/aoezkan/planeseg/eval/gt_moge_zp_benchmark/summary.csv"
DEFAULT_OUTPUT = "/cluster/scratch/aoezkan/planeseg/eval/gt_moge_zp_benchmark/summary_pretty.xlsx"


# ---------------------------------------------------------------------------
# Metric-direction classification
# ---------------------------------------------------------------------------
# ↑ = higher is better, ↓ = lower is better.

LOWER_IS_BETTER_PATTERNS: List[re.Pattern] = [
    re.compile(r"^VI$"),                          # Variation of Information
    re.compile(r"^DE_rel$"),
    re.compile(r"^DE_rel_sqr$"),
    re.compile(r"^DE_log10$"),
    re.compile(r"^DE_rmse$"),
    re.compile(r"^DE_rmse_log$"),
    re.compile(r"^pixel_DE_rel.*$"),              # if a separate pixel-depth pass exists
    re.compile(r"^pixel_DE_log10$"),
    re.compile(r"^pixel_DE_rmse.*$"),
    re.compile(r"^mean_normal_error_deg$"),
    re.compile(r"^median_normal_error_deg$"),
    re.compile(r"^mean_offset_error_m$"),
    re.compile(r"^median_offset_error_m$"),
    re.compile(r"^plane_param_L2_mean$"),
    re.compile(r"^plane_param_L2_area_weighted$"),
]

HIGHER_IS_BETTER_PATTERNS: List[re.Pattern] = [
    re.compile(r"^RI$"),
    re.compile(r"^SC$"),
    re.compile(r"^DE_accuracy_[123]$"),
    re.compile(r"^pixel_DE_accuracy_[123]$"),
    re.compile(r"^per_pixel_depth_.*$"),
    re.compile(r"^per_plane_depth_.*$"),
    re.compile(r"^per_pixel_normal_.*$"),
    re.compile(r"^per_plane_normal_.*$"),
    re.compile(r"^per_pixel_offset_.*$"),
    re.compile(r"^per_plane_offset_.*$"),
    re.compile(r"^AP@\d+cm$"),
    re.compile(r"^prec@.*cm$"),
    re.compile(r"^rec@.*cm$"),
]


def metric_direction(metric: str) -> Optional[str]:
    """Return '↑' / '↓' for a known metric, or None if unknown."""
    for p in LOWER_IS_BETTER_PATTERNS:
        if p.match(metric):
            return "↓"
    for p in HIGHER_IS_BETTER_PATTERNS:
        if p.match(metric):
            return "↑"
    return None


# ---------------------------------------------------------------------------
# Core
# ---------------------------------------------------------------------------

def _better(a, b, direction: str) -> Optional[str]:
    """Return 'a', 'b', or None (tie / NaNs)."""
    try:
        af, bf = float(a), float(b)
    except (TypeError, ValueError):
        return None
    if math.isnan(af) or math.isnan(bf) or af == bf:
        return None
    if direction == "↑":
        return "a" if af > bf else "b"
    if direction == "↓":
        return "a" if af < bf else "b"
    return None


def build_value_and_bold_frames(df: pd.DataFrame):
    """Return (values_df, bold_mask_df, header_renames).

    - ``values_df`` keeps the original numeric dtypes (good for Excel formatting).
    - ``bold_mask_df`` is the same shape, bool, True where the cell should
      render bold.
    - ``header_renames`` maps original columns to display names (with arrows).
    """
    if "method" not in df.columns or "dataset" not in df.columns:
        raise ValueError("input CSV must have 'method' and 'dataset' columns")

    mean_cols = [c for c in df.columns if c.endswith("_mean")]
    std_cols = [c for c in df.columns if c.endswith("_std")]
    keep_label_cols = [c for c in df.columns if c not in mean_cols + std_cols]

    direction_for: Dict[str, str] = {}     # original col -> '↑' / '↓' / ''
    rename_map: Dict[str, str] = {}
    for c in mean_cols:
        base = c[: -len("_mean")]
        d = metric_direction(base) or ""
        direction_for[c] = d
        rename_map[c] = f"{base} {d}".strip()
    for c in keep_label_cols:
        rename_map[c] = c

    out_cols = keep_label_cols + mean_cols
    values = df[out_cols].copy()
    bold_mask = pd.DataFrame(False, index=values.index, columns=out_cols)

    # Header (row 0 of openpyxl) is bolded uniformly. Method/dataset columns
    # are bolded for every row to anchor the eye.
    for c in keep_label_cols:
        if c in ("method", "dataset"):
            bold_mask[c] = True

    for i, row in df.iterrows():
        method = row["method"]
        dataset = row["dataset"]
        if method not in ("moge", "zeroplane"):
            continue
        other = "zeroplane" if method == "moge" else "moge"
        opp_rows = df[(df["dataset"] == dataset) & (df["method"] == other)]
        if len(opp_rows) != 1:
            continue
        opp = opp_rows.iloc[0]
        for c in mean_cols:
            d = direction_for[c]
            if not d:
                continue
            winner = _better(row[c], opp[c], d)
            if winner == "a":
                bold_mask.at[i, c] = True

    return values, bold_mask, rename_map


def write_xlsx(values: pd.DataFrame, bold_mask: pd.DataFrame,
               header_renames: Dict[str, str], out_path: str) -> None:
    """Write to xlsx with per-cell ``Font(bold=True)`` where mask is True."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise SystemExit(
            "openpyxl is required for xlsx output. Install with:\n"
            "    pip install openpyxl"
        ) from e

    wb = Workbook()
    ws = wb.active
    ws.title = "summary"

    bold_font = Font(bold=True)

    # Header row.
    display_cols = [header_renames.get(c, c) for c in values.columns]
    for j, name in enumerate(display_cols, start=1):
        cell = ws.cell(row=1, column=j, value=name)
        cell.font = bold_font

    # Data rows. We use ``enumerate(values.itertuples)`` so positional iteration
    # matches the bold_mask index.
    for i, (_, row) in enumerate(values.iterrows(), start=2):
        for j, c in enumerate(values.columns, start=1):
            v = row[c]
            # Cast NaN to empty so the cell is blank in Excel rather than "nan".
            if isinstance(v, float) and math.isnan(v):
                cell = ws.cell(row=i, column=j, value=None)
            else:
                cell = ws.cell(row=i, column=j, value=v)
            if bool(bold_mask.iat[i - 2, values.columns.get_loc(c)]):
                cell.font = bold_font

    # Number formatting for the mean columns: 4 significant figures-ish via
    # a general format, but with a sane fallback for very small numbers.
    num_fmt = "0.0000"
    for j, c in enumerate(values.columns, start=1):
        if not c.endswith("_mean"):
            continue
        col_letter = get_column_letter(j)
        for cell in ws[col_letter][1:]:    # skip header
            if isinstance(cell.value, (int, float)):
                cell.number_format = num_fmt

    # Freeze header row + the first two label columns (method, dataset) for
    # easier scrolling through the wide metric set.
    ws.freeze_panes = "C2"

    # Loose column-width auto-sizing based on the wider of header / sample value.
    for j, c in enumerate(values.columns, start=1):
        col_letter = get_column_letter(j)
        header_len = len(str(display_cols[j - 1]))
        sample = values[c].astype(str).str.len().max() if len(values) else 0
        try:
            sample_len = int(sample)
        except (TypeError, ValueError):
            sample_len = 0
        ws.column_dimensions[col_letter].width = min(40, max(8, header_len + 2, sample_len + 2))

    wb.save(out_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--input", default=DEFAULT_INPUT,
                    help=f"Path to summary.csv (default: {DEFAULT_INPUT})")
    ap.add_argument("--output", default=DEFAULT_OUTPUT,
                    help=f"Where to write the prettified xlsx "
                         f"(default: {DEFAULT_OUTPUT})")
    args = ap.parse_args()

    if not os.path.isfile(args.input):
        raise SystemExit(f"input not found: {args.input}")

    df = pd.read_csv(args.input)
    values, bold_mask, header_renames = build_value_and_bold_frames(df)
    write_xlsx(values, bold_mask, header_renames, args.output)
    print(f"wrote {args.output}  ({len(values)} rows × {len(values.columns)} cols)")

    # Report unknown metrics so direction-map gaps are visible.
    unknown = [c[:-len("_mean")] for c in df.columns
               if c.endswith("_mean") and metric_direction(c[:-len("_mean")]) is None]
    if unknown:
        print(f"  warning: no direction set for {len(unknown)} metric(s): {unknown}")


if __name__ == "__main__":
    main()
