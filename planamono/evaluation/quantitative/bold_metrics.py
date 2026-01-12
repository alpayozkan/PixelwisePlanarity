import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import os

# Input CSV files
precision_recall_file = 'table_precision_recall_baselines.csv'
segmentation_file = 'table_segmentation_baselines.csv'
output_file = 'table_combined_baselines_bold.xlsx'

# Get script directory for relative paths
script_dir = os.path.dirname(os.path.abspath(__file__))
precision_recall_path = os.path.join(script_dir, precision_recall_file)
segmentation_path = os.path.join(script_dir, segmentation_file)
output_path = os.path.join(script_dir, output_file)

# Load both CSVs
df_pr = pd.read_csv(precision_recall_path)
df_seg = pd.read_csv(segmentation_path)

# Merge on Method (keep num_scenes and num_frames from one)
df = pd.merge(
    df_seg[['Method', 'num_scenes', 'num_frames', 'RI', 'VOI', 'SC']],
    df_pr[['Method', 'P@0.1cm', 'R@0.1cm', 'P@0.5cm', 'R@0.5cm', 'P@1.0cm', 'R@1.0cm']],
    on='Method',
    how='outer'
)

# Reorder columns: Method, counts, segmentation metrics, then precision/recall
column_order = [
    'Method', 'num_scenes', 'num_frames',
    'SC', 'RI', 'VOI',
    'P@0.1cm', 'R@0.1cm', 'P@0.5cm', 'R@0.5cm', 'P@1.0cm', 'R@1.0cm'
]
df = df[[c for c in column_order if c in df.columns]]

# Round float columns to 3 decimal places
float_cols = df.select_dtypes(include=['float64']).columns
df[float_cols] = df[float_cols].round(3)

# Define method order: GT Seg, Ours, ZeroPlane, then ablations
method_order = [
    'GT Seg (upper bound)',
    'Ours (full)',
    'ZeroPlane',
    'Our Planarity + GT Seg',
    'GT Planarity + Our Seg',
]

# Sort by method order
df['sort_key'] = df['Method'].apply(lambda x: method_order.index(x) if x in method_order else len(method_order))
df = df.sort_values('sort_key').drop('sort_key', axis=1).reset_index(drop=True)

# Define metric types: True = higher is better, False = lower is better
metric_types = {
    'SC': True,       # Segmentation Covering - higher is better
    'RI': True,       # Rand Index - higher is better
    'VOI': False,     # Variation of Information - lower is better
    'P@0.1cm': True,
    'R@0.1cm': True,
    'P@0.5cm': True,
    'R@0.5cm': True,
    'P@1.0cm': True,
    'R@1.0cm': True,
}

# Find best values in DataFrame BEFORE saving (including all methods)
best_indices = {}
for col, higher_is_better in metric_types.items():
    if col not in df.columns:
        continue
    if higher_is_better:
        best_idx = df[col].idxmax()
    else:
        best_idx = df[col].idxmin()
    best_indices[col] = best_idx
    print(f"{col}: Best = {df.loc[best_idx, col]:.3f} ({df.loc[best_idx, 'Method']})")

# Save the dataframe first
df.to_excel(output_path, index=False)

# Load the workbook to apply formatting
wb = load_workbook(output_path)
ws = wb.active

# Get column indices (1-based for openpyxl)
header_row = [cell.value for cell in ws[1]]

# Apply bold to best values
for col, df_idx in best_indices.items():
    if col not in header_row:
        continue
    col_idx = header_row.index(col) + 1  # 1-based for openpyxl
    row_idx = df_idx + 2  # +2 because: +1 for header, +1 for 1-based indexing

    cell = ws.cell(row=row_idx, column=col_idx)
    cell.font = Font(bold=True)

# Save the workbook
wb.save(output_path)
print(f"\nSaved to {output_path}")
